#!/usr/bin/env python3
#-*- coding: utf-8 -*-

"""GHAS Reporting Tool

This script retrieves various types of GitHub Advanced Security (GHAS) alerts for a specified
organization or repository. The types of alerts include Code scanning alerts, Secret scanning
alerts, and Dependabot alerts. The script generates reports based on the specified options and
writes the results to a file in CSV, XLSX, JSON, or all formats.

The script uses the GitHub API to retrieve alert data and requires valid API credentials. 
An API key can be specified added using the "ghas_enc_key.py" script, or alternatively
specified in the GH_API_KEY environment variable.

Usage:
    python ghas_reporting_tool.py [options]

options:
  -h, --help            show this help message and exit
  -v, --version         show program's version number and exit

Generate alert reports:
  -a, --all             generate Alert Count, Code Scanning, Secret Scanning, and Dependabot alert reports
  -l, --alerts          generate Alert Count report of all open alerts
  -c, --codescan        generate Code Scan alert report
  -s, --secretscan      generate Secret Scanning alert report
  -d, --dependabot      generate Dependabot alert report

Optional alert state arguments:
  -o, --open            generate report(s) for open alerts only - note: this setting has no effect on the alert count report "--alerts", which only includes open alerts

Output file format arguments:
  -wA, --output-all     write output to all formats at once
  -wC, --output-csv     write output to a CSV file (default format)
  -wX, --output-xlsx    write output to a Microsoft Excel file
  -wJ, --output-json    write output to a JSON file

Optional file format arguments:
  -t <theme>, --theme <theme>
                        specify the color theme for "xlsx" file output. Valid keywords are "grey", "blue", "green", "rose", "purple", "aqua", "orange". If none is specified, defaults to "grey".

Optional alert report arguments:
  -n, --owner           specify the owner of a GitHub repository, or organization. required if the "--repo" or "--org" options are specified.
  -g, --org             specify the name of a GitHub organization. This option is mutually exclusive with the "--repo" option. The "--owner" option is required if this option is specified.
  -r, --repo            specify the name of a GitHub repository. This option is mutually exclusive with the "--org" option. The "--owner" option is required if this option is specified.

Optional location arguments:
  -lc, --config         specify file location for the configuration file ("ghas_conf.json")
  -lk, --keyfile        specify file location for the encryption key file (".ghas_env") - overrides the location specified in the configuration file
  -lr, --reports        specify file location for the reports directory - overrides the location specified in the configuration file

Requirements:
    - Python 3.6 or later
    - requests
    - cryptography
    - openpyxl

Dependencies:
    - ghas_enc_key.py

Package: ghas_report.py
Version: 1.2.0-dev
Date: XXXX-XX-XX

Author: Rupert Herbst <rhe8502(at)pm.me>
Project URL: https://github.com/rhe8502/ghas_report
License: Apache License, Version 2.0 (http://www.apache.org/licenses/LICENSE-2.0)
"""
# Copyright (C) 2023 Rupert Herbst
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.worksheet.hyperlink import Hyperlink
from cryptography.fernet import Fernet
from datetime import datetime
import argparse
import csv
import json
import requests
import os
import sys
import time
import re
import openpyxl

def api_error_response(response):
    """Generate error message from API response.

    Args:
        response (requests.Response): API response object.

    Returns:
        str: Error message.

    Examples:
        >>> response = requests.get('https://api.example.com')
        >>> api_error_response(response)
        'Error 404: Resource not found'
    """
    error_messages = {
        304: f"Error {response.status_code}: {response.json().get('message', 'Not modified')}" + (f", Errors: {response.json().get('errors', '')}" if response.json().get('errors') else ''),
        400: f"Error {response.status_code}: {response.json().get('message', 'Bad Request')}" + (f", Errors: {response.json().get('errors', '')}" if response.json().get('errors') else ''),
        403: f"Error {response.status_code}: {response.json().get('message', 'GitHub Advanced Security is not enabled for this repository')}" + (f", Errors: {response.json().get('errors', '')}" if response.json().get('errors') else ''),
        404: f"Error {response.status_code}: {response.json().get('message', 'Resource not found')}" + (f", Errors: {response.json().get('errors', '')}" if response.json().get('errors') else ''),
        422: f"Error {response.status_code}: {response.json().get('message', 'Validation failed, or the endpoint has been spammed')}" + (f", Errors: {response.json().get('errors', '')}" if response.json().get('errors') else ''),
        503: f"Error {response.status_code}: {response.json().get('message', 'Service unavailable')}" + (f", Errors: {response.json().get('errors', '')}" if response.json().get('errors') else '')
    }
    
    if response.status_code in error_messages:
        error_message = error_messages[response.status_code]
        raise Exception(error_message)
    else:
        raise Exception(f"Error {response.status_code}: {response.json().get('message', '')}" + (f", Errors: {response.json().get('errors', '')}" if response.json().get('errors') else ''))

def get_scan_alerts(api_url, org_name=None, call_func=None, owner=None, repo_name=None, state=None):
    """Retrieve alerts for a specific scan type from the GitHub API and count the open alerts and their corresponding severity levels.

    Args:
        api_url (str): The base URL for the GitHub API.
        org_name (str, optional): The name of the organization to retrieve alerts for.
        call_func (str, optional): The type of scan alerts to retrieve ('codescan', 'secretscan', 'dependabot').
        owner (str, optional): The name of the repository owner.
        repo_name (str, optional): The name of the repository to retrieve alerts for.
        state (str, optional): The state of the alerts to retrieve ('open', 'closed', or None for all).

    Returns:
        tuple: A tuple containing a list of scan_alerts and a list of severity counts (sev_list).

    Nested Functions:
        alerts_count(alerts, sev_counts)
        get_next_page_link(link_header)
    """
    scan_types = {
        'codescan': 'code-scanning',
        'secretscan': 'secret-scanning',
        'dependabot': 'dependabot'
    }

    base_url = f"{api_url}/repos/{owner}/{repo_name}/{scan_types[call_func]}/alerts" if repo_name else f"{api_url}/orgs/{org_name}/{scan_types[call_func]}/alerts"
    state_query = '&state=open' if state == 'open' else ''

    scan_alerts = []
    open_alert_count = 0
    sev_counts = {
        'critical': 0,
        'high': 0,
        'medium': 0,
        'low': 0,
        'warning': 0,
        'note': 0,
        'error': 0
    }

    def alerts_count(alerts, sev_counts):
        """Counts the open alerts and their corresponding severity levels for the given list of alerts.

        Args:
            alerts (list): A list of alert dictionaries.
            sev_counts (dict): A dictionary to store the counts for each severity level.

        Returns:
            None
        """
        nonlocal open_alert_count
        for alert in alerts:
            if alert['state'] == 'open':
                open_alert_count += 1
                sev = None

                if call_func == 'codescan':
                    sev = alert.get('rule', {}).get('security_severity_level') or alert.get('rule', {}).get('severity', '').lower()
                elif call_func == 'dependabot':
                    sev = alert['security_advisory']['severity'].lower() if 'severity' in alert['security_advisory'] else None

                if sev and sev in sev_counts:
                    sev_counts[sev] += 1

    def get_next_page_link(link_header):
        """Extracts the next page URL from the 'Link' header in a paginated API response.

        Args:
            link_header (str): The 'Link' header value from an API response.

        Returns:
            str: The URL of the next page, or None if not found.
        """
        if link_header:
            next_page_link = re.search(r'<(.+?)>; rel="next"', link_header)
            return next_page_link.group(1) if next_page_link else None
        return None

    if call_func == 'codescan':
        page_no = 1

        while True:
            url = f"{base_url}?page={page_no}{state_query}"
            response = requests.get(url, headers=headers)

            if response.status_code == 200:
                alerts = response.json()

                if not alerts:
                    break

                alerts_count(alerts, sev_counts)
                scan_alerts.extend(alerts)
                page_no += 1
            else:
                print(api_error_response(response))
                break

    elif call_func in ['secretscan', 'dependabot']:
        url = f"{base_url}?per_page=100{state_query}"

        while url:
            response = requests.get(url, headers=headers)

            if response.status_code == 200:
                alerts = response.json()

                if not alerts:
                    break

                alerts_count(alerts, sev_counts)
                scan_alerts.extend(alerts)

                # Get the next page URL from the 'Link' header if present
                link_header = response.headers.get('Link')
                url = get_next_page_link(link_header)
            else:
                print(api_error_response(response))
                break

    sev_list = [val for val in sev_counts.values()]
    sev_list.insert(0, open_alert_count)

    return scan_alerts, sev_list

def process_alerts_count(api_url, project_data):
    """Processes and retrieves the alert count for each scan type (Code Scan, Secret Scan, Dependabot Scan)
       for the specified organizations and repositories.

    Description:
        The function iterates through the organizations and repositories provided in the project_data dictionary and retrieves the alert count
        for each scan type (Code Scan, Secret Scan, Dependabot Scan). It then appends the alert count for each organization or repository along
        with the corresponding scan type to a list.

    Args:
        api_url (str): The API URL to retrieve scan data.
        project_data (dict): A dictionary containing project data, including organizations, repositories, and owner (if repositories are specified).

    Returns:
        dict: A dictionary containing the raw alert count data and the processed alert count data as lists.
    """
    alert_count = []

    scan_types = {
        'Code Scan': 'codescan',
        'Secret Scan': 'secretscan',
        'Dependabot Scan': 'dependabot'
    }

    for gh_entity in ['organizations', 'repositories']:
        for gh_name in project_data.get(gh_entity, []):
            if gh_name:
                try:
                    for scan_label, call_func in scan_types.items():
                        if gh_entity == 'organizations':
                            sev_list = get_scan_alerts(api_url, org_name=gh_name, call_func=call_func)[1]
                        elif gh_entity == 'repositories':
                            owner = project_data.get('owner')
                            sev_list = get_scan_alerts(api_url, owner=owner, repo_name=gh_name, call_func=call_func)[1]

                        row = [gh_name if gh_entity == 'organizations' else '', gh_name if gh_entity == 'repositories' else '', scan_label, *sev_list]
                        alert_count.append(row)
                except Exception as e:
                    print(f"Error getting alert count for {'repository' if gh_entity == 'repositories' else 'organization'}: {gh_name} - {e}")

    return {'raw_alerts': alert_count, 'scan_alerts': alert_count}

def get_theme(output_theme):
    # Grey light theme
    if output_theme == 'grey':
        header_fill = PatternFill(start_color="232323", end_color="232323", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        odd_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        data_font = Font(size=11)
        border_color = "BFBFBF"
        thin_border = Border(
            left=Side(style='thin', color=border_color),
            right=Side(style='thin', color=border_color),
            top=Side(style='thin', color=border_color),
            bottom=Side(style='thin', color=border_color)
        )
        cell_alignment = Alignment(vertical="center", wrap_text=False)
        hyperlink_style = Font(color="1C1C1C", size=11, underline="single")
        hyperlink_alignment = Alignment(vertical="center", wrap_text=False)

    # Blue light theme
    elif output_theme == 'blue':
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        odd_row_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        data_font = Font(size=11)
        border_color = "95B3D7"
        thin_border = Border(
            left=Side(style='thin', color=border_color),
            right=Side(style='thin', color=border_color),
            top=Side(style='thin', color=border_color),
            bottom=Side(style='thin', color=border_color)
        )
        cell_alignment = Alignment(vertical="center", wrap_text=False)
        hyperlink_style = Font(color="0000EE", size=11, underline="single")
        hyperlink_alignment = Alignment(vertical="center", wrap_text=False)

    # Rose light theme
    elif output_theme == 'rose':
        header_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        odd_row_fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        data_font = Font(size=11)
        border_color = "DA9694"
        thin_border = Border(
            left=Side(style='thin', color=border_color),
            right=Side(style='thin', color=border_color),
            top=Side(style='thin', color=border_color),
            bottom=Side(style='thin', color=border_color)
        )
        cell_alignment = Alignment(vertical="center", wrap_text=False)
        hyperlink_style = Font(color="0000EE", size=11, underline="single")
        hyperlink_alignment = Alignment(vertical="center", wrap_text=False)

    # Green light theme
    elif output_theme == 'green':
        header_fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        odd_row_fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
        even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        data_font = Font(size=11)
        border_color = "C4D79B"
        thin_border = Border(
            left=Side(style='thin', color=border_color),
            right=Side(style='thin', color=border_color),
            top=Side(style='thin', color=border_color),
            bottom=Side(style='thin', color=border_color)
        )
        cell_alignment = Alignment(vertical="center", wrap_text=False)
        hyperlink_style = Font(color="0000EE", size=11, underline="single")
        hyperlink_alignment = Alignment(vertical="center", wrap_text=False)

    # Purple light theme
    elif output_theme == 'purple':
        header_fill = PatternFill(start_color="8064A2", end_color="8064A2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        odd_row_fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
        even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        data_font = Font(size=11)
        border_color = "B1A0C7"
        thin_border = Border(
            left=Side(style='thin', color=border_color),
            right=Side(style='thin', color=border_color),
            top=Side(style='thin', color=border_color),
            bottom=Side(style='thin', color=border_color)
        )
        cell_alignment = Alignment(vertical="center", wrap_text=False)
        hyperlink_style = Font(color="0000EE", size=11, underline="single")
        hyperlink_alignment = Alignment(vertical="center", wrap_text=False)

    # Aqua light theme
    elif output_theme == 'aqua':
        header_fill = PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        odd_row_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        data_font = Font(size=11)
        border_color = "92CDDC"
        thin_border = Border(
            left=Side(style='thin', color=border_color),
            right=Side(style='thin', color=border_color),
            top=Side(style='thin', color=border_color),
            bottom=Side(style='thin', color=border_color)
        )
        cell_alignment = Alignment(vertical="center", wrap_text=False)
        hyperlink_style = Font(color="0000EE", size=11, underline="single")
        hyperlink_alignment = Alignment(vertical="center", wrap_text=False)

    # Orange light theme
    elif output_theme == 'orange':
        header_fill = PatternFill(start_color="F79646", end_color="F79646", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        odd_row_fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
        even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        data_font = Font(size=11)
        border_color = "FABF8F"
        thin_border = Border(
            left=Side(style='thin', color=border_color),
            right=Side(style='thin', color=border_color),
            top=Side(style='thin', color=border_color),
            bottom=Side(style='thin', color=border_color)
        )
        cell_alignment = Alignment(vertical="center", wrap_text=False)
        hyperlink_style = Font(color="0000EE", size=11, underline="single")
        hyperlink_alignment = Alignment(vertical="center", wrap_text=False)

    theme = {
        'header_fill': header_fill,
        'header_font': header_font,
        'odd_row_fill': odd_row_fill,
        'even_row_fill': even_row_fill,
        'data_font': data_font,
        'border_color': border_color,
        'thin_border': thin_border,
        'cell_alignment': cell_alignment,
        'hyperlink_style': hyperlink_style,
        'hyperlink_alignment': hyperlink_alignment
    }

    return theme

def write_xlsx(header_row, alert_data, project_name, filepath, call_func, output_theme=None):
    """Writes alert data to an XLSX file.

    This function writes the given alert data to an XLSX file. 

    Args:
    header_row (list): A list of column header labels for the XLSX file.
    alert_data (dict): A dictionary containing the alert data to be written to the XLSX file.
    project_name (str): The name of the project for which the alert data is being written.
    filepath (str): The full file path where the XLSX file will be saved.
    call_func (str): The function calling write_xlsx, used to set the sheet name and select appropriate headers for the output file.
    """
    # Call the get_theme function and store the returned theme dictionary
    theme = get_theme(output_theme)

    # Replace the theme-related variables with values from the theme dictionary
    header_fill = theme['header_fill']
    header_font = theme['header_font']
    odd_row_fill = theme['odd_row_fill']
    even_row_fill = theme['even_row_fill']
    data_font = theme['data_font']
    border_color = theme['border_color']
    thin_border = theme['thin_border']
    cell_alignment = theme['cell_alignment']
    hyperlink_style = theme['hyperlink_style']
    hyperlink_alignment = theme['hyperlink_alignment']

    try:
        wb = openpyxl.load_workbook(filepath)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    
    # Add a new worksheet with the name of the function calling write_xlsx
    ws = wb.create_sheet(call_func)

    # Remove the default "Sheet" if it exists
    if "Sheet" in wb.sheetnames:
        default_sheet = wb["Sheet"]
        wb.remove(default_sheet)

    for col_num, col_data in enumerate(header_row):
        cell = ws.cell(row=1, column=col_num + 1, value=col_data)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = cell_alignment

    # Set filter on the header row
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(header_row))}1"

    # Write the data rows and apply alternate row colors
    for row_num, row_data in enumerate(alert_data['scan_alerts'], start=2):
        row_fill = odd_row_fill if row_num % 2 == 0 else even_row_fill
        for col_num, col_data in enumerate(row_data):
            cell = ws.cell(row=row_num, column=col_num + 1, value=col_data)

            # Check if the cell value is zero, then set the data type to 'n' (number)
            if col_data == '0':
                cell.data_type = 'n'
    
            cell.fill = row_fill
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = cell_alignment

    # Set this flag to True if the current sheet contains URLs
    contains_urls = True  # Change this value based on your sheet's content

    if call_func != 'alert_count' and contains_urls:
        # Assuming the last column contains URLs, loop through the rows and add hyperlinks
        url_column = len(header_row)  # Change this value if the URL column is not the last one
        for row_num in range(2, len(alert_data['scan_alerts']) + 2):
            cell = ws.cell(row=row_num, column=url_column)
            url = cell.value
            if url:
                try:
                    # Add the friendly name as the URL itself
                    cell.value = f'=HYPERLINK("{url}", "{url}")'
                    cell.font = hyperlink_style
                    cell.alignment = hyperlink_alignment
                except Exception as e:
                        print(f"Error while processing hyperlink at row {row_num}: {e}")

    if call_func == 'alert_count':
        # Calculate the total for each column in the first worksheet
        first_ws = wb.worksheets[0]
        last_row = first_ws.max_row

        # Define cell formatting styles
        bold_font = Font(bold=True)
        border_style = Border(top=Side(style='medium',color=border_color), bottom=Side(style='medium', color=border_color))
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        first_ws.cell(row=last_row + 1, column=1, value="Total").font = bold_font

        # Loop through the columns and calculate the total
        for col_num in range(4, 12):  # Columns D-K (4-11)
            column_letter = openpyxl.utils.get_column_letter(col_num)
            column_total = sum(first_ws.cell(row=row_num, column=col_num).value for row_num in range(2, last_row + 1))
            cell = first_ws[f"{column_letter}{last_row + 1}"]
            cell.value = column_total
        
        for col_num in range (1, 12):
            cell = first_ws.cell(row=last_row + 1, column=col_num)
            cell.border = border_style
            cell.font = bold_font
            cell.fill = white_fill

    # Set the maximum cell length - default is 60
    max_cell_length = 60 
    
    #Adjust padding for column width, if needed - default is 0
    padding = 0
    
    # Autosize column width
    for col_num, col_data in enumerate(header_row, start=1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)

        for row_num in range(1, len(alert_data['scan_alerts']) + 3):  # +3 to include header and one extra row for safety
            cell_value = str(ws.cell(row=row_num, column=col_num).value)
            cell_length = len(cell_value)
            max_length = max(max_length, cell_length)

        # Limit the maximum column width to the specified maximum cell length
        max_length = min(max_length, max_cell_length)

        # Set the column width
        ws.column_dimensions[column_letter].width = max_length + padding

    # Add extra padding for the header row to compensate for the filter dropdown icon
    header_padding = 5  # Adjust this value as needed
    for col_num, col_data in enumerate(header_row, start=1):
        column_letter = openpyxl.utils.get_column_letter(col_num)
        ws.column_dimensions[column_letter].width += header_padding
    
    # Freeze top row
    ws.freeze_panes = "A2"

    # Save workbook
    wb.save(filepath)
    print(f"Wrote {call_func} for \"{project_name}\" to {filepath}")

def write_alerts(alert_data, project_name, output_type=None, output_theme=None, report_dir='', call_func=None, time_stamp=None):
    """Writes the processed scan alert data to a file in the specified format (CSV, XLSX, or JSON).

    Description:
        The function writes the processed scan alert data to a file in the specified output format (CSV, XLSX, or JSON).
        It sets the column headers for the CSV file depending on the type of alert. If the output type is not specified, it defaults to 'csv'.
        The function creates a file path based on the report directory, project name, alert type, and the current date and time.

    Args:
        alert_data (dict): A dictionary containing processed alert data.
        project_name (str): The name of the project.
        output_type (str, optional): The output format for the alert data file; either 'csv' or 'json'. Defaults to 'csv'.
        report_dir (str, optional): The directory where the report file should be saved. Defaults to an empty string, which means the file will be saved in a folder named after the current date.
        call_func (str, optional): The function to be called for processing alerts; either 'codescan', 'secretscan', or 'dependabot'. Defaults to None.

    Raises:
        SystemExit: If there's an error writing to the file.
    """
    # Set scan type to GHAS-Report if xlsx is defined, otherwise set it to the call_func
    scan_type = 'GHAS_Report' if output_type == 'xlsx' else call_func

    # Check if a report path is defined and create the file path, otherwise create a folder for the current date and create the file path
    if report_dir:
        filepath = os.path.join(report_dir, f"{project_name}-{scan_type}-{time_stamp}.{output_type}")
    else:
        filepath = os.path.join(report_dir, f"{datetime.now().strftime('%Y%m%d')}", f"{project_name}-{scan_type}-{time_stamp}.{output_type}")

    os.makedirs(os.path.dirname(filepath), exist_ok=True)
   
    # Set the column headers for the CSV file depending on the type of alert
    scan_options = {
        'alert_count': ['Organization', 'Repository', 'Scan Type', 'Total Alerts', 'Critical', 'High', 'Medium', 'Low', 'Warning', 'Note', 'Error'],
        'code_scan': ['Alert', 'Organization', 'Repository', 'Date Created', 'Date Updated', 'Days Open', 'Severity', 'State', 'Rule ID', 'Description', 'Category', 'File', 'Fixed At', 'Dismissed At', 'Dismissed By', 'Dismissed Reason', 'Dismissed Comment', 'Tool', 'GitHub URL'],
        'secret_scan': ['Alert', 'Organization', 'Repository', 'Date Created', 'Date Updated', 'Days Open', 'State', 'Resolved At', 'Resolved By', 'Resolved Reason', 'Secret Type Name', 'Secret Type', 'GitHub URL'],
        'dependabot_scan': ['Alert', 'Organization', 'Repository', 'Date Created', 'Date Updated', 'Days Open', 'Severity', 'State', 'Package Name', 'CVE ID', 'Summary', 'Fixed At', 'Dismissed At', 'Dismissed By', 'Dismissed Reason', 'Dismissed Comment', 'Scope', 'Manifest ID', 'GitHub URL']
    }

    # Set the header row
    header_row = scan_options.get(call_func, [])
    
    # Write the alert data to a file in the specified format, if none is specified, default to CSV
    if output_type == 'xlsx':
        write_xlsx(header_row, alert_data, project_name, filepath, call_func, output_theme)
    else:
        try:
            with open(filepath, 'w', encoding='utf-8', newline='') as f:
                writer = csv.writer(f) if output_type == 'csv' else None
                if output_type == 'json':
                    json.dump(alert_data['raw_alerts'], f, indent=4)
                    print(f"Wrote {call_func} for \"{project_name}\" to {filepath}")
                elif output_type == 'csv':
                    writer.writerow(header_row)
                    writer.writerows(alert_data['scan_alerts'])
                    print(f"Wrote {call_func} for \"{project_name}\" to {filepath}")
        except IOError as e:
            raise SystemExit(f"Error writing to {e.filename}: {e}")

def safe_get(alert, keys, default=''):
    """Safely retrieves the value from a nested dictionary using a list of keys.

    Args:
        alert (dict): The dictionary from which to retrieve the value.
        keys (list): A list of keys to traverse the dictionary.
        default (any, optional): The default value to return if any of the keys are not found. Defaults to an empty string.

    Returns:
        any: The value found in the dictionary using the provided keys, or the default value if any of the keys are not found.
    """
    result = alert
   
    for key in keys:
        if result:
            result = result.get(key)
        else:
            break
    
    return default if result is None else result

def process_scan_alerts(api_url, project_data, call_func, output_type=None ,state=None):
    """Retrieves and processes scan alerts from GitHub organizations and repositories.

    Description:
        The function iterates through each organization and repository, retrieves the corresponding scan alerts, processes them, and adds the processed alerts
        to a list. If output_type is set to 'json', the function returns raw alerts and skips further processing. The alerts are processed depending on the
        call_func parameter, which can be 'codescan', 'secretscan', or 'dependabot'. The function returns a dictionary containing the raw alerts and the processed
        scan alerts.

    Args:
        api_url (str): The base API URL for GitHub.
        project_data (dict): A dictionary containing project information such as owner, organizations, and repositories.
        call_func (str): The function to be called for processing alerts; either 'codescan', 'secretscan', or 'dependabot'.
        output_type (str, optional): The output type for raw alerts, defaults to None. If 'json', the function returns raw alerts and skips further processing.
        state (str, optional): Filter alerts based on their state, defaults to None.

    Returns:
        dict: A dictionary containing raw_alerts and processed scan_alerts.
    """
    raw_alerts = []
    scan_alerts = []

    # Iterate through each organization and repository and get the alerts
    for gh_entity in ['organizations', 'repositories']:
        for gh_name in project_data.get(gh_entity, []):
            if gh_name:
                try:
                    owner = project_data.get('owner') if gh_entity == 'repositories' else None
                    alerts = get_scan_alerts(api_url, owner=owner, org_name=gh_name, call_func=call_func, state=state, repo_name=gh_name if gh_entity == 'repositories' else None)[0]

                    # If the output type is json return the raw alerts and ignore the rest of the function
                    if output_type == 'json':
                        raw_alerts = alerts
                        return({'raw_alerts': raw_alerts})

                    # Process alerts for each organization and repository and add them to a list
                    for alert in alerts:
                        # Get the days open since the alert was created
                        days_since_created = (datetime.now() - datetime.strptime(safe_get(alert, ['created_at']), '%Y-%m-%dT%H:%M:%SZ')).days if safe_get(alert, ['created_at']) != '' else ''
                        
                        # Add default values for all alert types to the list
                        alert_data = [
                            safe_get(alert, ['number']),
                            gh_name if gh_entity == 'organizations' else safe_get(alert, ['organization', 'name'], ''),
                            gh_name if gh_entity == 'repositories' else safe_get(alert, ['repository', 'name'], ''),
                            datetime.strptime(safe_get(alert, ['created_at']), '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d') if safe_get(alert, ['created_at']) != '' else '',
                            datetime.strptime(safe_get(alert, ['updated_at']), '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d') if safe_get(alert, ['updated_at']) != '' else '',
                            days_since_created if safe_get(alert, ['state']) == 'open' else '0'
                        ]

                        # Add Code Scanning alert data to the list
                        if call_func == 'codescan':
                            alert_data.extend([
                                safe_get(alert, ['rule', 'security_severity_level']) or safe_get(alert, ['rule', 'severity'], ''),
                                safe_get(alert, ['state'], ''),
                                safe_get(alert, ['rule', 'id'], ''),
                                safe_get(alert, ['most_recent_instance', 'message', 'text'], ''),
                                safe_get(alert, ['most_recent_instance', 'category'], ''),
                                safe_get(alert, ['most_recent_instance', 'location', 'path'], ''),
                                datetime.strptime(safe_get(alert, ['fixed_at']), '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d') if safe_get(alert, ['fixed_at']) != '' else '',
                                datetime.strptime(safe_get(alert, ['dismissed_at']), '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d') if safe_get(alert, ['dismissed_at']) != '' else '',                               
                                safe_get(alert, ['dismissed_by', 'login'], ''),
                                safe_get(alert, ['dismissed_reason'], ''),
                                safe_get(alert, ['dismissed_comment'], ''),
                                safe_get(alert, ['tool', 'name'], '') + ' ' + safe_get(alert, ['tool', 'version'], ''),
                                safe_get(alert, ['html_url'], '')
                            ])
                                
                        # Add Secret Scanning alert data to the list
                        elif call_func == 'secretscan':
                            alert_data.extend([
                                safe_get(alert, ['state'], ''),
                                datetime.strptime(safe_get(alert, ['resolved_at']), '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d') if safe_get(alert, ['resolved_at']) != '' else '',
                                safe_get(alert, ['resolved_by', 'login'], ''),
                                safe_get(alert, ['resolution'], ''),
                                safe_get(alert, ['secret_type_display_name'], ''),
                                safe_get(alert, ['secret_type'], ''),
                                safe_get(alert, ['html_url'], '')
                            ])
                
                        # Add Dependabot alert data to the list
                        elif call_func == 'dependabot':
                            alert_data.extend([
                                safe_get(alert, ['security_advisory', 'severity'], ''),
                                safe_get(alert, ['state'], ''),
                                safe_get(alert, ['dependency', 'package', 'name'], ''),
                                safe_get(alert, ['security_advisory', 'cve_id'], ''),
                                safe_get(alert, ['security_advisory', 'summary'], ''),
                                datetime.strptime(safe_get(alert, ['fixed_at']), '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d') if safe_get(alert, ['fixed_at']) != '' else '',
                                datetime.strptime(safe_get(alert, ['dismissed_at']), '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d') if safe_get(alert, ['dismissed_at']) != '' else '',
                                safe_get(alert, ['dismissed_by', 'login'], ''),
                                safe_get(alert, ['dismissed_reason'], ''),
                                safe_get(alert, ['dismissed_comment'], ''),
                                safe_get(alert, ['dependency', 'scope'], ''),
                                safe_get(alert, ['dependency', 'manifest_path'], ''),
                                safe_get(alert, ['html_url'], '')
                            ])
                        scan_alerts.append(alert_data)
                except Exception as e:
                    print(f"Error getting {call_func} alerts for {'repository' if gh_entity == 'repositories' else 'organization'}: {gh_name} - {e}")
    
    return {'raw_alerts': raw_alerts, 'scan_alerts': scan_alerts}

def load_configuration(args):
    """Load the configuration and API key for making requests to the GitHub API.

    This function reads a configuration file and API key from the file system, optionally using
    command-line arguments to specify the file paths, and returns the parsed configuration and
    headers for API requests. If the API key is encrypted, it will be decrypted using the
    encryption key from the specified keyfile. If any required files are missing or the API key
    is corrupted, an error message will be displayed and the script will exit.

    Args:
    args (argparse.Namespace): The command-line arguments passed to the script.

    Returns:
    tuple: A tuple containing the loaded configuration (as a dictionary) and headers
    (as a dictionary) for making requests to the GitHub API.

    Raises:
    SystemExit: If the configuration file, keyfile, or API key is not found or if the
    API key is corrupted.
    """
    # Configuration file name and encryption key file name
    conf_file_name = 'ghas_config.json'
    env_file_name = '.ghas_env'
    config = {}

    # Determine script location and check if a commandline argument was passed for the config file, and if so, use that instead of the default
    script_dir = os.path.dirname(os.path.abspath(__file__))
    conf_file = os.path.join(args.config, conf_file_name) if args.config else os.path.join(script_dir, conf_file_name)

    # Get API key from environment variable
    api_key = os.environ.get("GH_API_KEY")

    # Load configuration file and get API key from it if not specified as an environment variable
    try:
        with open(conf_file) as f:
            config = json.load(f)
        if not api_key:
            api_key = config.get('connection', {}).get('gh_api_key', '')
    except FileNotFoundError as e:
        if not api_key:
            raise SystemExit(f"Error: No API key specified, or \"{conf_file}\" not found. Please run the \"ghas_enc_key.py\" script to add your API key.")
    except json.JSONDecodeError as e:
        raise SystemExit(f"Error loading {conf_file}: {e}\nYou might need to run the \"ghas_enc_key.py\" script to add your API key.")
       
    # if the API key is not specified as an environment variable, get the encryption key from the keyfile and decrypt the API key from the config file
    if not os.environ.get("GH_API_KEY"):
        enc_path = config.get('location', {}).get('keyfile','')

        # Check if a commandline argument was passed for the keyfile, if not, check if the keyfile path was specified in the config file, and if not, use the default
        if args.keyfile:
            env_file = os.path.join(args.keyfile, env_file_name)
        elif enc_path:
            env_file = os.path.join(enc_path, env_file_name)
        else:
            env_file = os.path.join(script_dir, env_file_name)  
        
        try:
            with open(env_file, 'rb') as f:
                f_key = f.read()
        except FileNotFoundError as e:
            raise SystemExit(f"Error loading {e.filename}: {e}\nYou might need to run the \"ghas_enc_key.py\" script first to generate a new \"{e.filename}\" file.")
        
        try:
            fernet = Fernet(f_key)
            api_key = fernet.decrypt(api_key.encode()).decode()
            config['connection']['gh_api_key'] = api_key
        except Exception :
            raise SystemExit(f"Error: Invalid key, your API key might be corrupted. Please run the \"ghas_enc_key.py\" script to encrypt the API key.") 

    # Define headers for API requests to GitHub
    headers = {
        "Authorization": f"token {api_key}",
        "X-GitHub-Api-Version": "2022-11-28" # API version number, see https://docs.github.com/en/rest/overview/api-versions
    }

    return config, headers

def setup_argparse():
    """Creates and returns an ArgumentParser object for the GitHub Advanced Security (GHAS) reporting tool.

    The ArgumentParser is configured with arguments for generating different types of alert reports, output formats, and file locations.

    Returns:
        argparse.ArgumentParser: The configured ArgumentParser object.
    """
    # Version number, release date, URL, license, and author
    version_number = '1.2.0-dev'
    release_date = 'XXXX-XX-XX'
    license = 'Apache License, Version 2.0 (http://www.apache.org/licenses/LICENSE-2.0)'
    url = 'https://github.com/rhe8502/ghas_report'
    author = "Rupert Herbst <rhe8502(at)pm.me>"

    # version string
    version_string = f"\n\nGitHub Advanced Security Reporting Tool v{version_number} ({release_date})\n\n{license}\nProject URL: {url}\n\nWritten by {author}"

    # Command-line arguments parser
    parser = argparse.ArgumentParser(description='''The script is designed to retrieve various types of GitHub Advanced Security (GHAS) alerts for a specified organization or repository. GHAS alerts can include Code scanning alerts, Secret scanning alerts, and Dependabot alerts.
                                                    \nThe script will generate a report based on the specified options and write the results to a file. The output format of the report can be specified using command-line options. The supported formats are CSV, XLSX and JSON. If no file format argument is specified then the output is written to a CSV file. If the -wA option is specified, then the report(s) will be written to all supported formats.''', formatter_class=argparse.RawTextHelpFormatter)

    # Options group
    parser.add_argument('-v', '--version', action='version', version=(version_string), help="show program's version number and exit")

    # Alert reports
    alert_group = parser.add_argument_group('Generate alert reports')
    alert_group.add_argument('-a', '--all', action='store_true', help='generate Alert Count, Code Scanning, Secret Scanning, and Dependabot alert reports')
    alert_group.add_argument('-l', '--alerts', action='store_true', help='generate Alert Count report of all open alerts')
    alert_group.add_argument('-c', '--codescan', action='store_true', help='generate Code Scan alert report')
    alert_group.add_argument('-s', '--secretscan', action='store_true', help='generate Secret Scanning alert report')
    alert_group.add_argument('-d', '--dependabot', action='store_true', help='generate Dependabot alert report')

    # Optional alert state arguments
    alert_state_group = parser.add_argument_group('Optional alert state arguments')
    alert_state_group.add_argument('-o', '--open', action='store_true', help='generate report(s) for open alerts only - note: this setting has no effect on the alert count report "--alerts", which only includes open alerts')

    # Output file format arguments
    output_group = parser.add_argument_group('Output file format arguments')
    output_group.add_argument('-wA', '--output-all', action='store_true', help='write output to all formats at once')
    output_group.add_argument('-wC', '--output-csv', action='store_true', help='write output to a CSV file (default format)')
    output_group.add_argument('-wX', '--output-xlsx', action='store_true', help='write output to a Microsoft Excel file')
    output_group.add_argument('-wJ', '--output-json', action='store_true', help='write output to a JSON file')
     
    # Optional file format arguments
    output_format_group = parser.add_argument_group('Optional file format arguments')
    output_format_group.add_argument('-t', '--theme', metavar='<theme>', type=str, choices=['grey', 'blue', 'rose', 'green', 'purple', 'aqua', 'orange'], default='grey', help='specify the color theme for "xlsx" file output. Valid keywords are "grey", "blue", "green", "rose", "purple", "aqua", "orange". If none is specified, defaults to "grey".')

    # Optional alert reports arguments
    alert_options_group = parser.add_argument_group('Optional alert report arguments')
    alert_options_group.add_argument('-n', '--owner', metavar='<owner>', type=str, help='specify the owner of a GitHub repository, or organization. required if the "--repo" or "--org" options are specified.')
    alert_options_group.add_argument('-g', '--org', metavar='<org>', type=str, help='specify the name of a GitHub organization. This option is mutually exclusive with the "--repo" option. The "--owner" option is required if this option is specified.')
    alert_options_group.add_argument('-r', '--repo', metavar='<repo>', type=str, help='specify the name of a GitHub repository. This option is mutually exclusive with the "--org" option. The "--owner" option is required if this option is specified.')

    # Optional location arguments
    location_options_group = parser.add_argument_group('Optional location arguments')
    location_options_group.add_argument('-lc', '--config', metavar='<PATH>', type=str, help='specify file location for the configuration file ("ghas_conf.json")')
    location_options_group.add_argument('-lk', '--keyfile', metavar='<PATH>', type=str, help='specify file location for the encryption key file (".ghas_env") - overrides the location specified in the configuration file')
    location_options_group.add_argument('-lr', '--reports', metavar='<PATH>', type=str, help='specify file location for the reports directory - overrides the location specified in the configuration file')

    return parser

def check_args_errors(args, parser):
    """Check for errors in the command-line arguments and display appropriate error messages.

    This function checks the given command-line arguments for inconsistencies or errors, such as
    missing required arguments, conflicting arguments, or no arguments at all. If any issues are
    detected, the function will display an error message and the script will exit.

    Args:
        args (argparse.Namespace): The command-line arguments passed to the script.
        parser (argparse.ArgumentParser): The ArgumentParser object used for parsing the command-line arguments.

    Raises:
        SystemExit: If any errors or inconsistencies are detected in the command-line arguments.
    """
    if len(sys.argv) == 1:
        parser.print_help()
        raise SystemExit('\nError: No arguments specified. Please specify at least one alert type --all, --alerts, --codescan, --secretscan, or --dependabot.\n')
    elif not any([args.all, args.alerts, args.codescan, args.secretscan, args.dependabot]):
        parser.print_help()
        raise SystemExit('\nError: No alert type specified. Please specify at least one alert type --all, --alerts, --codescan, --secretscan, or --dependabot.\n')
    elif args.output_all and (args.output_csv or args.output_json):
        parser.print_help()
        raise SystemExit('\nError: --output-all cannot be used together with --output-csv or --output-json\n')
    elif args.repo and args.org:
        parser.print_help()
        raise SystemExit('\nError: --repo and --org cannot be used together.\n')
    elif args.repo and not args.owner:
        parser.print_help()
        raise SystemExit('\nError: --repo requires --owner to be specified.\n')
    elif args.org and not args.owner:
        parser.print_help()
        raise SystemExit('\nError: --org requires --owner to be specified.\n')
    elif args.owner and not (args.repo or args.org):
        parser.print_help()
        raise SystemExit('\nError: --owner requires --repo or --org to be specified.\n')

def process_args(parser):
    """Process the command-line arguments and execute the appropriate functions.

    This function parses the command-line arguments, checks for errors, and based on the provided
    arguments, calls the necessary functions to generate alert reports in the specified output
    formats. It also loads the configuration file and handles the processing of the alert types
    and output types.

    Args:
        parser (argparse.ArgumentParser): The ArgumentParser object used for parsing the command-line arguments.

    Raises:
        SystemExit: If any errors or inconsistencies are detected in the command-line arguments.
    """
    args = parser.parse_args()

    # Call the check_args_errors function with the arguments and parser
    check_args_errors(args, parser)

    # Define the list of alert types to process. If the -a flag is present, include all alert types. Otherwise, include only the alert types that were passed as arguments
    alert_types = ['alerts', 'codescan', 'secretscan', 'dependabot'] if args.all else [alert_type for alert_type in ['alerts', 'codescan', 'secretscan', 'dependabot'] if getattr(args, alert_type)]
    
    # Define the list of output types to process. If the -wA flag is present, include all output types. Otherwise, include only the output types that were passed as arguments, if no output types are specified, default to CSV
    output_types = ['csv', 'xlsx', 'json'] if args.output_all else [output_type for output_type in ['csv', 'xlsx', 'json'] if getattr(args, f'output_{output_type}')] or ['csv']
    
    # Get the theme color from the --theme flag, or default to 'grey'
    output_theme = args.theme

    # Set state to 'open' if the -o ,or --open flag is present
    alert_state = 'open' if args.open else ''

    # Call the load_configuration function to load the configuration file and assign the returned values to the config and headers variables 
    global headers # Not too happy about this, but it's the only way I could get the headers variable to be accessible throughout the script
    config, headers = load_configuration(args)
    api_url = config.get('connection', {}).get('gh_api_url', '') or "https://api.github.com"
    report_dir = args.reports or config.get('location', {}).get('reports', '')
    time_stamp = datetime.now().strftime('%Y%m%d%H%M%S')
    
    # Check if the org or repo arguments are present. If they are, set use_config to False. Otherwise, set use_config to True
    use_config = not (args.org or args.repo)

    projects = (
        {args.org or args.owner: {key: value for key, value in [('owner', args.owner), ('organizations', [args.org]), ('repositories', [args.repo])] if value}}
        if not use_config
        else config.get('projects', {})
    )

    # Iterate through projects, alert types, and output types
    for project_name, project_data in projects.items():
        for alert_type in alert_types:
            for output_type in output_types:
                {
                    'alerts': lambda: write_alerts(process_alerts_count(api_url, project_data), project_name, output_type, output_theme, report_dir, call_func='alert_count', time_stamp=time_stamp),
                    'codescan': lambda output_type=output_type: write_alerts(process_scan_alerts(api_url, project_data, 'codescan', output_type, alert_state), project_name, output_type, output_theme, report_dir, call_func='code_scan', time_stamp=time_stamp),
                    'secretscan': lambda output_type=output_type: write_alerts(process_scan_alerts(api_url, project_data, 'secretscan', output_type, alert_state), project_name, output_type, output_theme, report_dir, call_func='secret_scan', time_stamp=time_stamp),
                    'dependabot': lambda output_type=output_type: write_alerts(process_scan_alerts(api_url, project_data, 'dependabot', output_type, alert_state), project_name, output_type, output_theme, report_dir, call_func='dependabot_scan', time_stamp=time_stamp),
                }[alert_type]()

def execution_time(start_time):
    """Prints the script's execution time."""
    end_time = time.perf_counter()
    elapsed_time = end_time - start_time
    min, sec = divmod(elapsed_time, 60)
    print(f"\nScript execution time: {elapsed_time:.2f} seconds\n") if elapsed_time < 60 else print(f"\nScript execution time: {int(min):02d}:{int(sec):02d} minutes\n")

def main():
    # Start the timer to measure the script's execution time
    start_time = time.perf_counter()

    # Call the setup_argparse function to configure the ArgumentParser object and assign the returned value to the parser variable, then call the process_args function to process the command-line arguments
    parser = setup_argparse()
    process_args(parser)
    
    # Call the execution_time function to print the script's execution time
    execution_time(start_time)

if __name__ == '__main__':
    main()